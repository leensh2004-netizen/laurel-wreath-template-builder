from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


@dataclass
class MappingItem:
    group_name: str
    account_name: str
    sheet_name: str
    cell: str


def clean_text(value) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\s+", " ", text)

    replacements = {
        "أ": "ا",
        "إ": "ا",
        "آ": "ا",
        "ة": "ه",
        "ى": "ي",
        "ـ": "",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    return text.strip()


def arabic_words(value) -> set:
    text = clean_text(value)

    stop_words = {
        "في", "من", "الى", "إلى", "علي", "على", "لدي", "لدى",
        "و", "او", "أو", "عن", "مع", "كل", "غير",
    }

    words = []

    for word in text.split():
        word = word.strip()

        if word.startswith("ال") and len(word) > 4:
            word = word[2:]

        if word.endswith("ات") and len(word) > 4:
            word = word[:-2]

        if word.endswith("ون") and len(word) > 4:
            word = word[:-2]

        if word.endswith("ين") and len(word) > 4:
            word = word[:-2]

        if len(word) > 2 and word not in stop_words:
            words.append(word)

    return set(words)


def clean_number(value) -> float:
    if value is None:
        return 0.0

    text = str(value).strip()

    if not text or text.lower() == "nan":
        return 0.0

    negative = text.startswith("(") and text.endswith(")")

    text = (
        text.replace(",", "")
        .replace("(", "")
        .replace(")", "")
        .replace("د.أ", "")
        .replace("دينار", "")
        .strip()
    )

    try:
        number = float(text)
    except ValueError:
        return 0.0

    return -number if negative else number


def reset_file_pointer(file_obj) -> None:
    try:
        file_obj.seek(0)
    except Exception:
        pass


def format_amount(value) -> str:
    try:
        number = float(value or 0)
    except (TypeError, ValueError):
        number = 0.0

    if number == int(number):
        return f"{int(number):,}"

    return f"{number:,.3f}"


def make_trial_balance_option(row) -> str:
    detail = str(row.get("detail_account_name", "") or "").strip()
    group = str(row.get("account_name", "") or "").strip()
    current = format_amount(row.get("current_amount", 0))
    previous = format_amount(row.get("previous_amount", 0))

    label_name = detail or group

    return f"{label_name} | {group} | Current: {current} | Previous: {previous}"


def read_mapping_excel(mapping_file) -> List[MappingItem]:
    """
    Reads the mapping Excel.

    Expected:
    - Column A = code like A1, A1.1, L3.1
    - Column B = Arabic account/group name
    - Rows without dot like A1, A2, L3 are parent groups
    - Rows with dot like A1.1, A1.2 are detailed accounts
    """
    reset_file_pointer(mapping_file)

    wb = load_workbook(mapping_file, data_only=True)
    ws = wb.active

    items: List[MappingItem] = []
    current_group: Optional[str] = None

    for row in range(2, ws.max_row + 1):
        code = str(ws.cell(row=row, column=1).value or "").strip()
        desc = str(ws.cell(row=row, column=2).value or "").strip()

        if not code or not desc:
            continue

        if code in ["A", "L"]:
            current_group = None
            continue

        if desc.replace("*", "").strip() == "":
            continue

        # Parent rows: A1, A2, L1, L3...
        if "." not in code:
            current_group = desc
            continue

        # Detail rows: A1.1, A1.2, L3.1...
        if current_group:
            items.append(
                MappingItem(
                    group_name=current_group,
                    account_name=desc,
                    sheet_name=ws.title,
                    cell=ws.cell(row=row, column=2).coordinate,
                )
            )

    return items


def read_trial_balance(trial_balance_file) -> pd.DataFrame:
    """
    Reads the trial balance.

    Expected columns:
    - المجموعة = account group
    - النهائي = current year amount
    - column before المجموعة = previous year amount
    - إسم الحساب = detailed account name
    """
    reset_file_pointer(trial_balance_file)

    raw = pd.read_excel(trial_balance_file, header=None)

    header_row_index = None
    group_col = None
    current_col = None
    detail_col = None

    for r in range(min(20, len(raw))):
        row_values = [clean_text(v) for v in raw.iloc[r].values]

        for c, value in enumerate(row_values):
            original_value = str(raw.iloc[r, c])

            if "المجموعه" in value or "المجموعة" in original_value:
                header_row_index = r
                group_col = c

            if "النهائي" in value:
                current_col = c

            if "اسم الحساب" in value or "إسم الحساب" in original_value:
                detail_col = c

    if header_row_index is None or group_col is None or current_col is None:
        return pd.DataFrame(
            columns=[
                "account_name",
                "account_name_clean",
                "detail_account_name",
                "current_amount",
                "previous_amount",
            ]
        )

    previous_col = group_col - 1

    rows = []

    for r in range(header_row_index + 1, len(raw)):
        account_group = str(raw.iloc[r, group_col] or "").strip()

        if not account_group or account_group.lower() == "nan":
            continue

        detail_name = ""
        if detail_col is not None:
            detail_name = str(raw.iloc[r, detail_col] or "").strip()

        current_amount = clean_number(raw.iloc[r, current_col])
        previous_amount = clean_number(raw.iloc[r, previous_col])

        rows.append(
            {
                "account_name": account_group,
                "account_name_clean": clean_text(account_group),
                "detail_account_name": detail_name,
                "current_amount": current_amount,
                "previous_amount": previous_amount,
            }
        )

    return pd.DataFrame(rows)


def find_best_match(
    account_name: str,
    trial_balance_df: pd.DataFrame,
    used_indexes: Optional[set] = None,
) -> Optional[pd.Series]:
    used_indexes = used_indexes or set()
    target = clean_text(account_name)

    if not target or trial_balance_df.empty:
        return None

    if "detail_account_name_clean" not in trial_balance_df.columns:
        trial_balance_df["detail_account_name_clean"] = trial_balance_df[
            "detail_account_name"
        ].apply(clean_text)

    # 1. Exact match with detailed account name
    exact_detail = trial_balance_df[
        (~trial_balance_df.index.isin(used_indexes))
        & (trial_balance_df["detail_account_name_clean"] == target)
    ]

    if not exact_detail.empty:
        return exact_detail.iloc[0]

    # 2. Exact match with group/account name
    exact_group = trial_balance_df[
        (~trial_balance_df.index.isin(used_indexes))
        & (trial_balance_df["account_name_clean"] == target)
    ]

    if not exact_group.empty:
        return exact_group.iloc[0]

    # 3. Contains match on detailed account name
    for idx, row in trial_balance_df.iterrows():
        if idx in used_indexes:
            continue

        detail = str(row.get("detail_account_name_clean", ""))

        if len(target) >= 6 and (target in detail or detail in target):
            return row

    # 4. Arabic word overlap
    target_words = arabic_words(target)

    best_row = None
    best_score = 0.0

    for idx, row in trial_balance_df.iterrows():
        if idx in used_indexes:
            continue

        detail_words = arabic_words(row.get("detail_account_name_clean", ""))

        if not target_words or not detail_words:
            continue

        overlap = len(target_words & detail_words)
        score = overlap / max(len(target_words), len(detail_words))

        if score > best_score:
            best_score = score
            best_row = row

    if best_score >= 0.50:
        return best_row

    return None


def build_mapping_preview(
    mapping_file,
    trial_balance_file,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    mapping_items = read_mapping_excel(mapping_file)
    trial_balance_df = read_trial_balance(trial_balance_file)

    matched_rows: List[Dict[str, object]] = []
    used_trial_indexes = set()

    for item in mapping_items:
        match = find_best_match(
            item.account_name,
            trial_balance_df,
            used_indexes=used_trial_indexes,
        )

        if match is None:
            matched_rows.append(
                {
                    "Group": item.group_name,
                    "Mapping account": item.account_name,
                    "Suggested trial balance account": "",
                    "User selected trial balance account": "",
                    "Matched trial balance account": "",
                    "Current amount": 0.0,
                    "Previous amount": 0.0,
                    "Status": "Needs review",
                    "Mapping cell": item.cell,
                }
            )
        else:
            used_trial_indexes.add(match.name)

            match_label = make_trial_balance_option(match)

            matched_rows.append(
                {
                    "Group": item.group_name,
                    "Mapping account": item.account_name,
                    "Suggested trial balance account": match_label,
                    "User selected trial balance account": match_label,
                    "Matched trial balance account": match.get(
                        "detail_account_name",
                        match["account_name"],
                    ),
                    "Current amount": match["current_amount"],
                    "Previous amount": match["previous_amount"],
                    "Status": "Auto matched",
                    "Mapping cell": item.cell,
                }
            )

    preview_df = pd.DataFrame(matched_rows)

    summary_df = build_summary(preview_df)

    return preview_df, summary_df


def build_summary(preview_df: pd.DataFrame) -> pd.DataFrame:
    if preview_df.empty:
        return pd.DataFrame(
            columns=["Group", "Current total", "Previous total", "Matched accounts"]
        )

    valid_df = preview_df[preview_df["Status"] != "Needs review"]

    if valid_df.empty:
        return pd.DataFrame(
            columns=["Group", "Current total", "Previous total", "Matched accounts"]
        )

    summary_df = (
        valid_df.groupby("Group", as_index=False)
        .agg(
            {
                "Current amount": "sum",
                "Previous amount": "sum",
                "Mapping account": "count",
            }
        )
        .rename(
            columns={
                "Current amount": "Current total",
                "Previous amount": "Previous total",
                "Mapping account": "Matched accounts",
            }
        )
    )

    return summary_df


def get_trial_balance_options(trial_balance_file) -> List[str]:
    trial_balance_df = read_trial_balance(trial_balance_file)

    if trial_balance_df.empty:
        return [""]

    options = trial_balance_df.apply(make_trial_balance_option, axis=1).tolist()

    return [""] + options


def apply_reviewed_matches(
    reviewed_df: pd.DataFrame,
    trial_balance_file,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    trial_balance_df = read_trial_balance(trial_balance_file)

    if trial_balance_df.empty:
        return reviewed_df, pd.DataFrame(
            columns=["Group", "Current total", "Previous total", "Matched accounts"]
        )

    trial_balance_df = trial_balance_df.copy()
    trial_balance_df["trial_balance_option"] = trial_balance_df.apply(
        make_trial_balance_option,
        axis=1,
    )

    lookup = {
        row["trial_balance_option"]: row
        for _, row in trial_balance_df.iterrows()
    }

    updated_rows = []

    for _, row in reviewed_df.iterrows():
        selected = str(row.get("User selected trial balance account", "") or "").strip()

        new_row = row.to_dict()

        if selected and selected in lookup:
            match = lookup[selected]

            new_row["Matched trial balance account"] = match.get(
                "detail_account_name",
                match["account_name"],
            )
            new_row["Current amount"] = match["current_amount"]
            new_row["Previous amount"] = match["previous_amount"]
            new_row["Status"] = "Reviewed"
        else:
            new_row["Matched trial balance account"] = ""
            new_row["Current amount"] = 0.0
            new_row["Previous amount"] = 0.0
            new_row["Status"] = "Needs review"

        updated_rows.append(new_row)

    updated_df = pd.DataFrame(updated_rows)
    summary_df = build_summary(updated_df)

    return updated_df, summary_df
